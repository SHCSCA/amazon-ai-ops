import argparse
import csv
import hashlib
import itertools
import json
import os
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

REAL_REPORT_EXTENSIONS = {".xlsx", ".xls", ".csv"}
IMPORTABLE_BATCH_STATUSES = {"completed", "completed_with_errors"}
EVIDENCE_FILE_NAME_PATTERN = re.compile(
    r"(manifest|audit|diagnostic|screenshot|dom|trace|evidence|acceptance|batch-result|downloaded-report-files|failure)",
    re.IGNORECASE,
)
FIELD_MAPPING = {
    "日期": "date",
    "日期范围": "date",
    "数据日期": "date",
    "店铺": "storeName",
    "店铺名称": "storeName",
    "店铺名": "storeName",
    "站点": "marketplaceCode",
    "站点/市场": "marketplaceCode",
    "国家": "marketplaceCode",
    "广告组合": "portfolioName",
    "ASIN": "asin",
    "asin": "asin",
    "Asin": "asin",
    "MSKU": "msku",
    "msku": "msku",
    "Msku": "msku",
    "SKU": "sku",
    "sku": "sku",
    "广告活动": "campaignName",
    "广告活动名称": "campaignName",
    "Campaign": "campaignName",
    "campaign name": "campaignName",
    "广告组": "adGroupName",
    "广告组名称": "adGroupName",
    "Ad Group": "adGroupName",
    "ad group": "adGroupName",
    "关键词": "targeting",
    "关键词/ASIN": "targeting",
    "投放关键词": "targeting",
    "投放": "targeting",
    "Target": "targeting",
    "target": "targeting",
    "Search Term": "searchTerm",
    "search term": "searchTerm",
    "搜索词": "searchTerm",
    "用户搜索词": "searchTerm",
    "匹配方式": "matchType",
    "match type": "matchType",
    "Match Type": "matchType",
    "匹配类型": "matchType",
    "展现量": "impressions",
    "展示量": "impressions",
    "曝光量": "impressions",
    "Impressions": "impressions",
    "impressions": "impressions",
    "点击量": "clicks",
    "点击": "clicks",
    "Clicks": "clicks",
    "clicks": "clicks",
    "花费": "cost",
    "花费-本币": "cost",
    "花费金额": "cost",
    "消耗": "cost",
    "Cost": "cost",
    "cost": "cost",
    "Spend": "cost",
    "spend": "cost",
    "订单数": "orders",
    "订单": "orders",
    "广告订单": "orders",
    "Orders": "orders",
    "orders": "orders",
    "转化数": "orders",
    "销售额": "sales",
    "广告销售额-本币": "sales",
    "销售": "sales",
    "Sales": "sales",
    "sales": "sales",
    "GMV": "sales",
    "Revenue": "sales",
    "ACOS": "acos",
    "acos": "acos",
    "Acos": "acos",
    "CPC": "cpc",
    "CPC-本币": "cpc",
    "cpc": "cpc",
    "平均点击成本": "cpc",
    "点击成本": "cpc",
    "转化率": "cvr",
    "CVR": "cvr",
    "cvr": "cvr",
}


def normalize_key(value):
    key = str(value or "").strip()
    return FIELD_MAPPING.get(key) or FIELD_MAPPING.get(key.lower()) or key


def to_number(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    normalized = re.sub(r"[$,%\s,]", "", str(value)).strip()
    if normalized in {"", "--"}:
        return 0.0
    try:
        return float(normalized)
    except ValueError:
        return 0.0


def parse_date(value):
    if value is None or value == "":
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    match = re.match(r"(\d{4}-\d{2}-\d{2})", text)
    if match:
        return match.group(1)
    return text[:10]


def metric_from_row(row, batch, file_row):
    cost = to_number(row.get("cost"))
    clicks = int(round(to_number(row.get("clicks"))))
    orders = int(round(to_number(row.get("orders"))))
    sales = to_number(row.get("sales"))
    metric = {
        "batch_id": batch["id"],
        "report_type": file_row["report_type"],
        "portfolio_name": str(row.get("portfolioName") or ""),
        "date": parse_date(row.get("date")),
        "store_name": batch["store_name"] or str(row.get("storeName") or ""),
        "marketplace_code": batch["marketplace_code"] or str(row.get("marketplaceCode") or ""),
        "asin": str(row.get("asin") or ""),
        "msku": str(row.get("msku") or ""),
        "campaign_name": str(row.get("campaignName") or ""),
        "ad_group_name": str(row.get("adGroupName") or ""),
        "targeting": str(row.get("targeting") or ""),
        "search_term": str(row.get("searchTerm") or row.get("targeting") or ""),
        "match_type": str(row.get("matchType") or "exact"),
        "impressions": int(round(to_number(row.get("impressions")))),
        "clicks": clicks,
        "cost": cost,
        "orders": orders,
        "sales": sales,
        "currency": "USD",
        "acos": cost / sales if sales > 0 else 0,
        "cpc": cost / clicks if clicks > 0 else 0,
        "cvr": orders / clicks if clicks > 0 else 0,
        "source_file": file_row["file_path"],
        "source_row": row.get("__source_row"),
    }
    if not metric["date"]:
        return None
    if not any(metric[key] for key in ["asin", "campaign_name", "ad_group_name", "targeting", "search_term"]):
        return None
    return metric


def parse_workbook_rows(rows, file_row, batch):
    header_values = next(rows, None)
    if not header_values:
        return []
    headers = [normalize_key(value) for value in header_values]
    metrics = []
    for source_row, values in rows:
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
        row["__source_row"] = source_row
        metric = metric_from_row(row, batch, file_row)
        if metric:
            metrics.append(metric)
    return metrics


def parse_excel(file_row, batch):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("openpyxl is required to import .xlsx/.xls reports") from error

    workbook = load_workbook(file_row["file_path"], read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    iterator = worksheet.iter_rows(values_only=True)
    header_values = next(iterator, None)
    if not header_values:
        workbook.close()
        return []
    rows = enumerate(iterator, start=2)
    metrics = parse_workbook_rows(itertools.chain([header_values], rows), file_row, batch)
    workbook.close()
    return metrics


def parse_csv_file(file_row, batch):
    with open(file_row["file_path"], "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        header_values = next(reader, None)
        if not header_values:
            return []
        rows = enumerate(reader, start=2)
        return parse_workbook_rows(itertools.chain([header_values], rows), file_row, batch)


def parse_report_file(file_row, batch):
    suffix = Path(file_row["file_path"]).suffix.lower()
    if suffix == ".csv":
        return parse_csv_file(file_row, batch)
    return parse_excel(file_row, batch)


def sha256_file(file_path):
    digest = hashlib.sha256()
    with open(file_path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def ensure_column(conn, table, column, definition):
    columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def table_exists(conn, table):
    return conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table,),
    ).fetchone() is not None


def is_real_file(file_row, batch):
    file_path = file_row["file_path"]
    if file_row["status"] not in {"downloaded", "imported", "import_failed"} or not file_path:
        return False
    resolved = Path(file_path).resolve()
    download_dir = Path(batch["download_dir"]).resolve()
    if resolved.suffix.lower() not in REAL_REPORT_EXTENSIONS:
        return False
    if EVIDENCE_FILE_NAME_PATTERN.search(resolved.name):
        return False
    if not resolved.is_file():
        return False
    if resolved.stat().st_size <= 0:
        return False
    try:
      resolved.relative_to(download_dir)
    except ValueError:
      return False
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", required=True)
    parser.add_argument("--batch", required=True)
    parser.add_argument("--out")
    args = parser.parse_args()

    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row
    with conn:
        ensure_column(conn, "ad_daily_metrics", "currency", "TEXT DEFAULT 'USD'")
        ensure_column(conn, "ad_daily_metrics", "source_row", "INTEGER")
        if table_exists(conn, "report_files"):
            ensure_column(conn, "report_files", "file_hash", "TEXT")
            ensure_column(conn, "report_files", "import_error", "TEXT")
            ensure_column(conn, "report_files", "last_imported_at", "TEXT")
    batch_row = conn.execute("SELECT * FROM lingxing_report_batches WHERE id = ?", (args.batch,)).fetchone()
    if not batch_row:
        raise SystemExit(f"Batch not found: {args.batch}")
    batch = dict(batch_row)
    if batch["status"] not in IMPORTABLE_BATCH_STATUSES:
        raise SystemExit(f"Batch is not importable: {batch['status']}")
    file_rows = [
        dict(row)
        for row in conn.execute(
            "SELECT * FROM lingxing_report_files WHERE batch_id = ? ORDER BY report_type",
            (args.batch,),
        ).fetchall()
    ]
    real_files = [row for row in file_rows if is_real_file(row, batch)]
    if not real_files:
        raise SystemExit("No real .xlsx/.xls/.csv files found for batch")

    parse_results = []
    errors = []
    metrics = []
    imported_rows_by_file = {}
    errors_by_file = {}
    for file_row in real_files:
        try:
            parsed = parse_report_file(file_row, batch)
            imported_rows_by_file[file_row["file_path"]] = len(parsed)
            parse_results.append({
                "reportType": file_row["report_type"],
                "filePath": file_row["file_path"],
                "rows": len(parsed),
            })
            metrics.extend(parsed)
        except Exception as error:
            errors_by_file[file_row["file_path"]] = str(error)
            errors.append({
                "reportType": file_row["report_type"],
                "filePath": file_row["file_path"],
                "message": str(error),
            })

    with conn:
        deleted = conn.execute("DELETE FROM ad_daily_metrics WHERE batch_id = ?", (args.batch,)).rowcount
        conn.executemany(
            """
            INSERT INTO ad_daily_metrics (
              batch_id, report_type, portfolio_name,
              date, store_name, marketplace_code, asin, msku,
              campaign_name, ad_group_name, targeting, search_term, match_type,
              impressions, clicks, cost, orders, sales, currency, acos, cpc, cvr, source_file, source_row
            ) VALUES (
              :batch_id, :report_type, :portfolio_name,
              :date, :store_name, :marketplace_code, :asin, :msku,
              :campaign_name, :ad_group_name, :targeting, :search_term, :match_type,
              :impressions, :clicks, :cost, :orders, :sales, :currency, :acos, :cpc, :cvr, :source_file, :source_row
            )
            """,
            metrics,
        )
        if table_exists(conn, "report_files"):
            for file_row in real_files:
                file_path = file_row["file_path"]
                row_count = imported_rows_by_file.get(file_path, 0)
                import_error = errors_by_file.get(file_path)
                status = "import_failed" if import_error else "imported" if row_count > 0 else "downloaded"
                conn.execute(
                    """
                    INSERT INTO report_files (
                      batch_id, report_type, file_path, file_name, file_size, status,
                      imported_rows, file_hash, import_error, last_imported_at, created_at, updated_at
                    ) VALUES (
                      :batch_id, :report_type, :file_path, :file_name, :file_size, :status,
                      :imported_rows, :file_hash, :import_error, :last_imported_at, datetime('now'), datetime('now')
                    )
                    ON CONFLICT(batch_id, report_type, file_path) DO UPDATE SET
                      file_name = excluded.file_name,
                      file_size = excluded.file_size,
                      status = excluded.status,
                      imported_rows = excluded.imported_rows,
                      file_hash = excluded.file_hash,
                      import_error = excluded.import_error,
                      last_imported_at = excluded.last_imported_at,
                      updated_at = excluded.updated_at
                    """,
                    {
                        "batch_id": args.batch,
                        "report_type": file_row["report_type"],
                        "file_path": file_path,
                        "file_name": Path(file_path).name,
                        "file_size": Path(file_path).stat().st_size,
                        "status": status,
                        "imported_rows": row_count,
                        "file_hash": sha256_file(file_path),
                        "import_error": import_error,
                        "last_imported_at": datetime.now(timezone.utc).isoformat()
                        if row_count > 0 or import_error
                        else None,
                    },
                )

    totals = dict(conn.execute(
        """
        SELECT COUNT(*) AS rows, ROUND(SUM(cost), 2) AS spend, ROUND(SUM(sales), 2) AS sales,
               SUM(orders) AS orders, SUM(clicks) AS clicks
        FROM ad_daily_metrics
        WHERE batch_id = ?
        """,
        (args.batch,),
    ).fetchone())
    by_type = [
        dict(row)
        for row in conn.execute(
            """
            SELECT report_type AS reportType, COUNT(*) AS rows, ROUND(SUM(cost), 2) AS spend,
                   ROUND(SUM(sales), 2) AS sales, SUM(orders) AS orders, SUM(clicks) AS clicks
            FROM ad_daily_metrics
            WHERE batch_id = ?
            GROUP BY report_type
            ORDER BY report_type
            """,
            (args.batch,),
        ).fetchall()
    ]
    conn.close()

    evidence = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "dbPath": str(Path(args.db).resolve()),
        "batchId": args.batch,
        "scope": {
            "dateStart": batch["date_start"],
            "dateEnd": batch["date_end"],
            "storeName": batch["store_name"],
            "marketplaceCode": batch["marketplace_code"],
            "currency": "USD",
        },
        "downloadDir": batch["download_dir"],
        "manifestPath": batch["manifest_path"],
        "realFileCount": len(real_files),
        "parseResults": parse_results,
        "errors": errors,
        "deletedExisting": deleted,
        "inserted": len(metrics),
        "totals": totals,
        "byType": by_type,
    }
    if args.out:
        out_path = Path(args.out).resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(evidence, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "batchId": args.batch,
        "realFileCount": len(real_files),
        "inserted": len(metrics),
        "errors": len(errors),
        "totals": totals,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
